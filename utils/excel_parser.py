from openpyxl import load_workbook
from typing import List, Dict
from utils.logger import setup_logger

logger = setup_logger(__name__)

class ExcelParser:
    def __init__(self, file_path: str):
        self.file_path = file_path

    def parse(self) -> List[Dict]:
        """Parse Excel file and return test cases."""
        test_cases: List[Dict] = []

        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        # Read header row
        headers = [cell.value for cell in sheet[1]]
        logger.info(f"Headers: {headers}")

        # Process data rows
        for row in sheet.iter_rows(min_row=2, values_only=False):
            data: Dict[str, Any] = {}
            for idx, header in enumerate(headers):
                data[header] = row[idx].value

            # Accept row if it has either new or old URL headers
            url_a = data.get("URL_A") or data.get("UAT_URL")
            url_b = data.get("URL_B") or data.get("PROD_URL")

            if url_a and url_b:
                test_cases.append(data)
                logger.info(f"Loaded test case: {data.get('TEST_NAME') or url_a}")

        logger.info(f"Total test cases loaded: {len(test_cases)}")
        return test_cases
